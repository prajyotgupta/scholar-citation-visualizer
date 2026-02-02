"""
Two-step citation flow:
1) Fetch Aishwarya's top papers and generate per-paper XLSX tabs with inferred cities.
2) Read edited XLSX and generate a citation map, warning on unmapped cities.
"""

import argparse
import json
import math
import re
import time
from collections import defaultdict
from datetime import datetime

from scholarly import ProxyGenerator, scholarly
from geopy.geocoders import Nominatim
import plotly.graph_objects as go

AUTHOR_ID = "FA9h3ngAAAAJ"
DEFAULT_MAX_PAPERS = 4
DEFAULT_XLSX = "citations_review.xlsx"
DEFAULT_MAP = "citations_map_from_xlsx.png"
DEFAULT_UNMAPPED = "unmapped_cities.txt"
DEFAULT_COLUMN_F_XLSX = "citations_data.xlsx"
DEFAULT_COLUMN_F_OUTPUT = "aish_citation_world_map.png"
DEFAULT_CITY_CACHE = "city_mapping_cache.json"

# Light-weight institution-to-city mappings to improve city inference.
INSTITUTION_CITY_MAPPINGS = {
    "UCLA": "Los Angeles, California, USA",
    "Loyola University Chicago": "Chicago, Illinois, USA",
    "Washington State University": "Pullman, Washington, USA",
    "Harbin Institute of Technology": "Harbin, China",
    "University of Wisconsin-Madison": "Madison, Wisconsin, USA",
    "Tsinghua University": "Beijing, China",
    "Stanford University": "Stanford, California, USA",
    "New Mexico State University": "Las Cruces, New Mexico, USA",
    "WorldServe Education": "Bangalore, India",
    "COMSATS University": "Islamabad, Pakistan",
    "Macquarie University": "Sydney, Australia",
    "Lancaster University": "Lancaster, UK",
    "University of Houston": "Houston, Texas, USA",
    "University of Science and Technology of China": "Hefei, China",
    "Beijing Jiaotong University": "Beijing, China",
    "National Institute of Technology Hamirpur": "Hamirpur, India",
    "VNR VJIET": "Hyderabad, India",
    "SNS College of Technology": "Coimbatore, India",
    "Intel Corporation": "Santa Clara, California, USA",
    "Georgia Tech": "Atlanta, Georgia, USA",
    "Georgia Institute of Technology": "Atlanta, Georgia, USA",
    "GA Tech": "Atlanta, Georgia, USA",
}


def setup_scholarly():
    """Setup scholarly with proxy where possible to reduce rate limiting."""
    try:
        pg = ProxyGenerator()
        success = pg.FreeProxies()
        if success:
            scholarly.use_proxy(pg)
            print("✅ Using free proxy")
        else:
            print("⚠️  No proxy available, using direct connection")
    except Exception as exc:
        print(f"⚠️  Proxy setup failed: {exc}, using direct connection")


def retry_request(func, max_retries=3, delay=2):
    """Retry a function with exponential backoff."""
    for attempt in range(max_retries):
        try:
            return func()
        except Exception as exc:
            if attempt < max_retries - 1:
                wait_time = delay * (2 ** attempt)
                print(f"   ⚠️  Attempt {attempt + 1} failed: {exc}. Retrying in {wait_time}s...")
                time.sleep(wait_time)
            else:
                raise exc


def sanitize_sheet_title(title, existing_titles):
    """Ensure sheet title is <= 31 chars and unique."""
    cleaned = re.sub(r"[\[\]\*:/\\\?]", "", title).strip()
    if not cleaned:
        cleaned = "Paper"
    base = cleaned[:31]
    candidate = base
    counter = 2
    while candidate in existing_titles:
        suffix = f" ({counter})"
        candidate = (base[: 31 - len(suffix)] + suffix)[:31]
        counter += 1
    return candidate


def get_geocoder():
    return Nominatim(user_agent="citation_flow_geocoder", timeout=10)


def geocode_with_retries(geolocator, query, max_retries=3):
    for attempt in range(max_retries):
        try:
            location = geolocator.geocode(query)
            if location:
                return location
        except Exception as exc:
            print(f"   ⚠️  Geocode failed ({attempt + 1}/{max_retries}) for '{query}': {exc}")
        time.sleep(1.0)
    return None


def format_location_label(location):
    address = location.raw.get("address", {})
    city = (
        address.get("city")
        or address.get("town")
        or address.get("village")
        or address.get("municipality")
        or address.get("county")
    )
    state = address.get("state")
    country = address.get("country")

    if city and state and country:
        return f"{city}, {state}, {country}"
    if city and country:
        return f"{city}, {country}"
    if state and country:
        return f"{state}, {country}"
    if city:
        return city
    if country:
        return country
    return None


def looks_like_institution(value):
    keywords = [
        "university",
        "institute",
        "college",
        "school",
        "department",
        "laboratory",
        "centre",
        "center",
        "tech",
        "corporation",
        "inc",
        "ltd",
    ]
    lower = value.lower()
    return any(keyword in lower for keyword in keywords)


def load_city_cache(cache_path):
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {}


def save_city_cache(cache_path, cache):
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2, ensure_ascii=False)


def update_city_cache(raw_values, geolocator, cache_path):
    cache = load_city_cache(cache_path)
    updated = False

    for raw in sorted(raw_values):
        if raw in cache:
            continue

        query = INSTITUTION_CITY_MAPPINGS.get(raw, raw)
        if looks_like_institution(raw):
            query = INSTITUTION_CITY_MAPPINGS.get(raw, raw)

        location = geocode_with_retries(geolocator, query)
        if location:
            label = format_location_label(location) or query
            cache[raw] = {
                "label": label,
                "lat": location.latitude,
                "lon": location.longitude,
                "query": query,
            }
        else:
            cache[raw] = {"label": raw, "lat": None, "lon": None, "query": query}
        updated = True
        time.sleep(0.2)

    if updated:
        save_city_cache(cache_path, cache)

    return cache


def infer_city_from_affiliation(affiliation, geolocator):
    """Infer city string from an affiliation using mappings + geocoding."""
    if not affiliation:
        return None

    for institution, city in INSTITUTION_CITY_MAPPINGS.items():
        if institution.lower() in affiliation.lower():
            return city

    location = geocode_with_retries(geolocator, affiliation)
    if not location:
        return None

    address = location.raw.get("address", {})
    city = (
        address.get("city")
        or address.get("town")
        or address.get("village")
        or address.get("municipality")
        or address.get("county")
        or address.get("state")
    )
    country = address.get("country")

    if city and country:
        return f"{city}, {country}"
    if city:
        return city
    if country:
        return country
    return None


def fetch_author_with_publications(author_id):
    print("\n📚 Fetching author profile...")
    author = retry_request(lambda: scholarly.search_author_id(author_id))
    scholarly.fill(author, sections=["publications"])
    return author


def select_top_publications(publications, max_papers):
    """Select top publications by citation count."""
    ranked = sorted(
        publications,
        key=lambda p: p.get("num_citations", 0),
        reverse=True,
    )
    return ranked[:max_papers]


def fetch_citations_for_publication(publication):
    try:
        return list(scholarly.citedby(publication))
    except Exception as exc:
        print(f"   ❌ Error fetching citations: {exc}")
        return []


def build_citing_rows_for_publication(publication, geolocator):
    pub_title = publication.get("bib", {}).get("title", "Unknown Title")
    num_citations = publication.get("num_citations", 0)
    print(f"\n📄 {pub_title}")
    print(f"   Citations: {num_citations}")

    if num_citations == 0:
        return []

    citations = fetch_citations_for_publication(publication)
    print(f"   ✅ Retrieved {len(citations)} citing papers")

    rows = []
    for cit_idx, citation in enumerate(citations, 1):
        cit_title = citation.get("bib", {}).get("title", "Unknown")
        author_ids = citation.get("author_id", [])
        author_names = citation.get("bib", {}).get("author", [])
        if isinstance(author_names, str):
            author_names = [name.strip() for name in author_names.split(",")]

        authors = []
        cities = []

        for i, author_id in enumerate(author_ids):
            author_name = author_names[i] if i < len(author_names) else "Unknown"
            affiliation = None

            if author_id:
                try:
                    author_profile = scholarly.search_author_id(author_id)
                    author_name = author_profile.get("name", author_name)
                    affiliation = author_profile.get("affiliation", None)
                    time.sleep(0.3)
                except Exception as exc:
                    print(f"      ⚠️  Could not fetch profile for {author_name}: {exc}")

            authors.append(author_name)

            inferred_city = infer_city_from_affiliation(affiliation, geolocator)
            if inferred_city and inferred_city not in cities:
                cities.append(inferred_city)

        rows.append(
            {
                "citing_paper": cit_title,
                "authors": authors,
                "cities": cities,
            }
        )

        if cit_idx % 25 == 0:
            print(f"   ...processed {cit_idx}/{len(citations)} citations")

    return rows


def generate_review_xlsx(author_id, max_papers, output_xlsx):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("⚠️  openpyxl not installed. Installing...")
        import subprocess

        subprocess.check_call(["pip", "install", "openpyxl"])
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

    setup_scholarly()
    author = fetch_author_with_publications(author_id)
    publications = author.get("publications", [])
    selected_pubs = select_top_publications(publications, max_papers)

    geolocator = get_geocoder()

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    existing_titles = set()
    summary = []

    for pub in selected_pubs:
        pub_title = pub.get("bib", {}).get("title", "Unknown Title")
        sheet_title = sanitize_sheet_title(pub_title, existing_titles)
        existing_titles.add(sheet_title)

        rows = build_citing_rows_for_publication(pub, geolocator)
        summary.append({"paper": pub_title, "citations": len(rows)})

        ws = wb.create_sheet(title=sheet_title)

        max_city_cols = max((len(r["cities"]) for r in rows), default=0)
        headers = ["S.No", "Citing Paper", "Author Names"]
        headers.extend([f"City {i}" for i in range(1, max_city_cols + 1)])

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
            ws.cell(row=row_idx, column=2, value=row["citing_paper"]).border = thin_border
            ws.cell(row=row_idx, column=3, value="; ".join(row["authors"])).border = thin_border

            for city_idx, city in enumerate(row["cities"], 1):
                ws.cell(row=row_idx, column=3 + city_idx, value=city).border = thin_border

        # Adjust column widths
        column_widths = [8, 70, 45] + [30] * max_city_cols
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"

    wb.save(output_xlsx)
    print(f"\n✅ Review XLSX saved: {output_xlsx}")

    summary_path = "citations_review_summary.json"
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "author": author.get("name", "N/A"),
                "author_id": author_id,
                "generated_at": datetime.now().isoformat(),
                "papers": summary,
            },
            f,
            indent=2,
        )
    print(f"✅ Summary saved: {summary_path}")


def collect_cities_from_xlsx(xlsx_path):
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    city_counts = defaultdict(int)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue

        # Columns: 1 S.No, 2 Citing Paper, 3 Author Names, 4+ City columns
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            for cell in row[3:]:
                if cell and str(cell).strip():
                    city = str(cell).strip()
                    city_counts[city] += 1

    return city_counts


def generate_map_from_xlsx(xlsx_path, output_path, unmapped_path):
    city_counts = collect_cities_from_xlsx(xlsx_path)
    if not city_counts:
        print("❌ No cities found in XLSX. Check your data.")
        return

    geolocator = get_geocoder()
    latitudes = []
    longitudes = []
    labels = []
    sizes = []
    unmapped = []

    for city, count in sorted(city_counts.items(), key=lambda x: x[1], reverse=True):
        location = geocode_with_retries(geolocator, city)
        if not location:
            unmapped.append(city)
            continue
        latitudes.append(location.latitude)
        longitudes.append(location.longitude)
        labels.append(f"{city} ({count})")
        sizes.append(6 + 3 * math.log(count + 1))
        time.sleep(0.2)

    if not latitudes:
        print("❌ No cities could be geocoded. Map not generated.")
        return

    fig = go.Figure(
        data=go.Scattergeo(
            lon=longitudes,
            lat=latitudes,
            text=labels,
            mode="markers",
            marker=dict(size=sizes, color="blue", opacity=0.7),
        )
    )
    fig.update_geos(visible=True, scope="world", showcountries=True, countrycolor="Grey")
    fig.update_layout(title="Citations Map (from XLSX)")
    fig.write_image(output_path, scale=2)
    print(f"✅ Map saved: {output_path}")

    if unmapped:
        with open(unmapped_path, "w", encoding="utf-8") as f:
            f.write("\n".join(unmapped))
        print(f"⚠️  Unmapped cities ({len(unmapped)}). See: {unmapped_path}")
    else:
        print("✅ All cities mapped successfully.")


def collect_cities_from_column(xlsx_path, column_index):
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    city_counts = defaultdict(int)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < column_index:
                continue
            cell = row[column_index - 1]
            if not cell:
                continue
            raw = str(cell).strip()
            if not raw:
                continue
            values = [v.strip() for v in re.split(r"[;\n]", raw) if v.strip()]
            for value in values:
                city_counts[value] += 1

    return city_counts


def generate_map_from_column_f(xlsx_path, output_path, unmapped_path, cache_path):
    city_counts = collect_cities_from_column(xlsx_path, column_index=6)
    if not city_counts:
        print("❌ No cities found in Column F. Check your data.")
        return

    geolocator = get_geocoder()
    cache = update_city_cache(city_counts.keys(), geolocator, cache_path)

    latitudes = []
    longitudes = []
    labels = []
    sizes = []
    unmapped = []
    coords_by_label = {}
    resolved_counts = defaultdict(int)

    for raw_city, count in sorted(city_counts.items(), key=lambda x: x[1], reverse=True):
        entry = cache.get(raw_city, {})
        label = entry.get("label", raw_city)
        lat = entry.get("lat")
        lon = entry.get("lon")

        if lat is None or lon is None:
            location = geocode_with_retries(geolocator, label)
            if location:
                lat = location.latitude
                lon = location.longitude
                cache[raw_city] = {
                    "label": format_location_label(location) or label,
                    "lat": lat,
                    "lon": lon,
                    "query": entry.get("query", label),
                }
                save_city_cache(cache_path, cache)
            else:
                unmapped.append(raw_city)
                continue

        if label not in coords_by_label:
            coords_by_label[label] = (lat, lon)
        resolved_counts[label] += count

    if not resolved_counts:
        print("❌ No cities could be geocoded. Map not generated.")
        return

    labels = list(resolved_counts.keys())
    sizes = [resolved_counts[label] for label in labels]
    latitudes = [coords_by_label[label][0] for label in labels]
    longitudes = [coords_by_label[label][1] for label in labels]
    marker_sizes = [6 + 3 * math.log(size + 1) for size in sizes]
    marker_labels = [f"{label} ({size})" for label, size in zip(labels, sizes)]

    fig = go.Figure(
        data=go.Scattergeo(
            lon=longitudes,
            lat=latitudes,
            text=marker_labels,
            mode="markers",
            marker=dict(size=marker_sizes, color="blue", opacity=0.7),
        )
    )
    fig.update_geos(visible=True, scope="world", showcountries=True, countrycolor="Grey")
    fig.update_layout(title="Aishwarya Lekshmi Chithra's Citation Map")
    fig.write_image(output_path, scale=2)
    print(f"✅ Map saved: {output_path}")

    if unmapped:
        with open(unmapped_path, "w", encoding="utf-8") as f:
            f.write("\n".join(sorted(set(unmapped))))
        print(f"⚠️  Unmapped cities ({len(set(unmapped))}). See: {unmapped_path}")
    else:
        print("✅ All cities mapped successfully.")


def main():
    parser = argparse.ArgumentParser(description="Two-step citation flow.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    fetch_parser = subparsers.add_parser("fetch", help="Fetch citations and build review XLSX")
    fetch_parser.add_argument("--author-id", default=AUTHOR_ID, help="Google Scholar author ID")
    fetch_parser.add_argument("--max-papers", type=int, default=DEFAULT_MAX_PAPERS)
    fetch_parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Output XLSX path")

    map_parser = subparsers.add_parser("map", help="Generate map from edited XLSX")
    map_parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Input XLSX path")
    map_parser.add_argument("--output", default=DEFAULT_MAP, help="Output map image path")
    map_parser.add_argument("--unmapped", default=DEFAULT_UNMAPPED, help="Unmapped cities output")

    column_parser = subparsers.add_parser("map-column-f", help="Generate map from Column F")
    column_parser.add_argument("--xlsx", default=DEFAULT_COLUMN_F_XLSX, help="Input XLSX path")
    column_parser.add_argument("--output", default=DEFAULT_COLUMN_F_OUTPUT, help="Output map image path")
    column_parser.add_argument("--unmapped", default=DEFAULT_UNMAPPED, help="Unmapped cities output")
    column_parser.add_argument("--cache", default=DEFAULT_CITY_CACHE, help="City mapping cache JSON")

    args = parser.parse_args()

    if args.command == "fetch":
        generate_review_xlsx(args.author_id, args.max_papers, args.xlsx)
    elif args.command == "map":
        generate_map_from_xlsx(args.xlsx, args.output, args.unmapped)
    elif args.command == "map-column-f":
        generate_map_from_column_f(args.xlsx, args.output, args.unmapped, args.cache)


if __name__ == "__main__":
    main()

