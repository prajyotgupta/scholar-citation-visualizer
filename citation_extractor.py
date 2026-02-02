"""
Citation Extractor for Google Scholar
Generates XLSX and HTML outputs with citing authors' information
"""

from scholarly import scholarly, ProxyGenerator
import time
import json
from datetime import datetime

# Google Scholar Author ID
AUTHOR_ID = 'FA9h3ngAAAAJ'

def setup_scholarly():
    """Setup scholarly with retry mechanism"""
    # Try using free proxies to avoid rate limiting
    try:
        pg = ProxyGenerator()
        # Use free proxies (slower but avoids blocks)
        success = pg.FreeProxies()
        if success:
            scholarly.use_proxy(pg)
            print("✅ Using free proxy")
        else:
            print("⚠️  No proxy available, using direct connection")
    except Exception as e:
        print(f"⚠️  Proxy setup failed: {e}, using direct connection")

def retry_request(func, max_retries=3, delay=2):
    """Retry a function with exponential backoff"""
    for attempt in range(max_retries):
        try:
            return func()
        except Exception as e:
            if attempt < max_retries - 1:
                wait_time = delay * (2 ** attempt)
                print(f"   ⚠️  Attempt {attempt + 1} failed: {e}. Retrying in {wait_time}s...")
                time.sleep(wait_time)
            else:
                raise e

def get_scholar_profile_url(scholar_id):
    """Generate Google Scholar profile URL from scholar ID"""
    if scholar_id:
        return f"https://scholar.google.com/citations?user={scholar_id}"
    return None

def extract_email_domain(author_data):
    """Try to extract email domain from author data"""
    # Google Scholar shows "Verified email at domain.edu" in some profiles
    if 'email_domain' in author_data:
        return author_data['email_domain']
    return None

def process_all_citations():
    """
    Process all of Aishwarya's papers and extract citing authors' information.
    Returns structured data for XLSX and HTML generation.
    """
    print("="*60)
    print("CITATION EXTRACTOR")
    print("="*60)
    
    # Setup scholarly with proxy if needed
    setup_scholarly()
    
    # Search for author with retry
    print("\n📚 Fetching author profile...")
    try:
        author = retry_request(lambda: scholarly.search_author_id(AUTHOR_ID))
        scholarly.fill(author, sections=['publications'])
    except Exception as e:
        print(f"❌ Failed to fetch author: {e}")
        print("   Try running again in a few minutes (Google may be rate limiting)")
        return [], {}
    
    print(f"Author: {author.get('name', 'N/A')}")
    print(f"Total publications: {len(author.get('publications', []))}")
    
    # Data structure to hold all results
    all_data = []
    
    # Process each publication
    publications = author.get('publications', [])
    
    for pub_idx, pub in enumerate(publications, 1):
        pub_title = pub.get('bib', {}).get('title', 'Unknown Title')
        num_citations = pub.get('num_citations', 0)
        
        print(f"\n{'─'*60}")
        print(f"📄 [{pub_idx}/{len(publications)}] {pub_title}")
        print(f"   Citations: {num_citations}")
        
        if num_citations == 0:
            print("   ⏭️  No citations, skipping...")
            continue
        
        # Get citations for this publication
        try:
            citations = list(scholarly.citedby(pub))
            print(f"   ✅ Retrieved {len(citations)} citing papers")
        except Exception as e:
            print(f"   ❌ Error fetching citations: {e}")
            continue
        
        # Process each citation
        for cit_idx, citation in enumerate(citations, 1):
            cit_title = citation.get('bib', {}).get('title', 'Unknown')
            cit_year = citation.get('bib', {}).get('pub_year', 'N/A')
            
            print(f"\n   📝 Citation {cit_idx}/{len(citations)}: {cit_title[:50]}...")
            
            # Get author IDs and names
            author_ids = citation.get('author_id', [])
            author_names = citation.get('bib', {}).get('author', [])
            
            # Handle author names format
            if isinstance(author_names, str):
                author_names = [name.strip() for name in author_names.split(',')]
            
            # Process each author
            for i, author_id in enumerate(author_ids):
                author_name = author_names[i] if i < len(author_names) else "Unknown"
                
                author_info = {
                    'aish_paper': pub_title,
                    'aish_paper_year': pub.get('bib', {}).get('pub_year', 'N/A'),
                    'citing_paper': cit_title,
                    'citing_paper_year': cit_year,
                    'author_name': author_name,
                    'affiliation': None,
                    'email_domain': None,
                    'scholar_id': author_id,
                    'scholar_url': get_scholar_profile_url(author_id),
                    'has_profile': bool(author_id)
                }
                
                # If author has a Scholar ID, fetch their profile
                if author_id:
                    try:
                        author_profile = scholarly.search_author_id(author_id)
                        author_info['author_name'] = author_profile.get('name', author_name)
                        author_info['affiliation'] = author_profile.get('affiliation', None)
                        author_info['email_domain'] = author_profile.get('email_domain', None)
                        print(f"      ✓ {author_info['author_name']} | {author_info['affiliation'] or 'No affiliation'}")
                        time.sleep(0.3)  # Small delay to avoid rate limiting
                    except Exception as e:
                        print(f"      ⚠️  Could not fetch profile for {author_name}: {e}")
                else:
                    print(f"      ○ {author_name} (no profile)")
                
                all_data.append(author_info)
        
        # Save progress after each paper
        save_progress(all_data)
    
    return all_data, author

def save_progress(data):
    """Save progress to a JSON file in case of interruption"""
    with open('citation_data_progress.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def generate_xlsx(data, author_name):
    """Generate Excel file with citation data"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("⚠️  openpyxl not installed. Installing...")
        import subprocess
        subprocess.check_call(['pip', 'install', 'openpyxl'])
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Citations"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Aishwarya's Paper",
        "Citing Paper",
        "Author Name",
        "Affiliation/Designation",
        "Email Domain",
        "Scholar Profile"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Data rows
    for row_idx, item in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=item['aish_paper']).border = thin_border
        ws.cell(row=row_idx, column=2, value=item['citing_paper']).border = thin_border
        ws.cell(row=row_idx, column=3, value=item['author_name']).border = thin_border
        ws.cell(row=row_idx, column=4, value=item['affiliation'] or 'N/A').border = thin_border
        ws.cell(row=row_idx, column=5, value=item['email_domain'] or 'N/A').border = thin_border
        
        # Add hyperlink for scholar profile
        if item['scholar_url']:
            cell = ws.cell(row=row_idx, column=6, value="View Profile")
            cell.hyperlink = item['scholar_url']
            cell.font = link_font
        else:
            cell = ws.cell(row=row_idx, column=6, value="No Profile")
        cell.border = thin_border
    
    # Adjust column widths
    column_widths = [50, 50, 25, 40, 20, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    filename = 'citations_data.xlsx'
    wb.save(filename)
    print(f"\n✅ Excel file saved: {filename}")
    return filename

def generate_html(data, author_info):
    """Generate HTML page with citation data"""
    
    # Group data by Aishwarya's paper
    papers_dict = {}
    for item in data:
        paper = item['aish_paper']
        if paper not in papers_dict:
            papers_dict[paper] = {}
        
        citing_paper = item['citing_paper']
        if citing_paper not in papers_dict[paper]:
            papers_dict[paper][citing_paper] = []
        
        papers_dict[paper][citing_paper].append(item)
    
    # Count totals
    total_authors = len(data)
    total_citing_papers = sum(len(cites) for cites in papers_dict.values())
    authors_with_profiles = sum(1 for item in data if item['has_profile'])
    
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Citation Network - {author_info.get('name', 'Scholar')}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        :root {{
            --bg-primary: #0f172a;
            --bg-secondary: #1e293b;
            --bg-card: #334155;
            --text-primary: #f1f5f9;
            --text-secondary: #94a3b8;
            --accent: #38bdf8;
            --accent-hover: #0ea5e9;
            --success: #22c55e;
            --border: #475569;
        }}
        
        body {{
            font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
            background: var(--bg-primary);
            color: var(--text-primary);
            line-height: 1.6;
            min-height: 100vh;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            padding: 2rem;
        }}
        
        header {{
            text-align: center;
            padding: 3rem 0;
            border-bottom: 1px solid var(--border);
            margin-bottom: 2rem;
        }}
        
        h1 {{
            font-size: 2.5rem;
            background: linear-gradient(135deg, var(--accent), #818cf8);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            margin-bottom: 0.5rem;
        }}
        
        .subtitle {{
            color: var(--text-secondary);
            font-size: 1.1rem;
        }}
        
        .stats {{
            display: flex;
            justify-content: center;
            gap: 3rem;
            margin-top: 2rem;
            flex-wrap: wrap;
        }}
        
        .stat {{
            text-align: center;
        }}
        
        .stat-value {{
            font-size: 2.5rem;
            font-weight: bold;
            color: var(--accent);
        }}
        
        .stat-label {{
            color: var(--text-secondary);
            font-size: 0.9rem;
            text-transform: uppercase;
            letter-spacing: 1px;
        }}
        
        .search-box {{
            margin: 2rem 0;
            display: flex;
            justify-content: center;
        }}
        
        .search-box input {{
            width: 100%;
            max-width: 500px;
            padding: 1rem 1.5rem;
            border: 2px solid var(--border);
            border-radius: 50px;
            background: var(--bg-secondary);
            color: var(--text-primary);
            font-size: 1rem;
            outline: none;
            transition: border-color 0.3s;
        }}
        
        .search-box input:focus {{
            border-color: var(--accent);
        }}
        
        .search-box input::placeholder {{
            color: var(--text-secondary);
        }}
        
        .paper-section {{
            background: var(--bg-secondary);
            border-radius: 16px;
            margin-bottom: 1.5rem;
            overflow: hidden;
            border: 1px solid var(--border);
        }}
        
        .paper-header {{
            padding: 1.5rem;
            cursor: pointer;
            display: flex;
            justify-content: space-between;
            align-items: center;
            transition: background 0.3s;
        }}
        
        .paper-header:hover {{
            background: var(--bg-card);
        }}
        
        .paper-title {{
            font-size: 1.1rem;
            font-weight: 600;
            flex: 1;
        }}
        
        .paper-badge {{
            background: var(--accent);
            color: var(--bg-primary);
            padding: 0.25rem 0.75rem;
            border-radius: 20px;
            font-size: 0.85rem;
            font-weight: 600;
            margin-left: 1rem;
        }}
        
        .paper-content {{
            display: none;
            padding: 0 1.5rem 1.5rem;
        }}
        
        .paper-section.open .paper-content {{
            display: block;
        }}
        
        .paper-section.open .toggle-icon {{
            transform: rotate(180deg);
        }}
        
        .toggle-icon {{
            transition: transform 0.3s;
            color: var(--text-secondary);
        }}
        
        .citing-paper {{
            background: var(--bg-card);
            border-radius: 12px;
            padding: 1rem;
            margin-bottom: 1rem;
        }}
        
        .citing-title {{
            font-weight: 600;
            margin-bottom: 0.75rem;
            color: var(--text-primary);
            font-size: 0.95rem;
        }}
        
        .authors-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
            gap: 0.75rem;
        }}
        
        .author-card {{
            background: var(--bg-secondary);
            border-radius: 8px;
            padding: 0.75rem 1rem;
            border: 1px solid var(--border);
            transition: transform 0.2s, border-color 0.2s;
        }}
        
        .author-card:hover {{
            transform: translateY(-2px);
            border-color: var(--accent);
        }}
        
        .author-name {{
            font-weight: 600;
            color: var(--text-primary);
            display: flex;
            align-items: center;
            gap: 0.5rem;
        }}
        
        .author-name a {{
            color: var(--accent);
            text-decoration: none;
        }}
        
        .author-name a:hover {{
            text-decoration: underline;
        }}
        
        .profile-icon {{
            width: 16px;
            height: 16px;
            fill: var(--success);
        }}
        
        .author-affiliation {{
            color: var(--text-secondary);
            font-size: 0.85rem;
            margin-top: 0.25rem;
        }}
        
        .author-email {{
            color: var(--accent);
            font-size: 0.8rem;
            margin-top: 0.25rem;
        }}
        
        .no-profile {{
            opacity: 0.7;
        }}
        
        footer {{
            text-align: center;
            padding: 2rem;
            color: var(--text-secondary);
            border-top: 1px solid var(--border);
            margin-top: 2rem;
        }}
        
        .hidden {{
            display: none !important;
        }}
        
        @media (max-width: 768px) {{
            .container {{
                padding: 1rem;
            }}
            
            h1 {{
                font-size: 1.75rem;
            }}
            
            .stats {{
                gap: 1.5rem;
            }}
            
            .stat-value {{
                font-size: 2rem;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>📚 Citation Network</h1>
            <p class="subtitle">Papers citing {author_info.get('name', 'Scholar')}'s research</p>
            <div class="stats">
                <div class="stat">
                    <div class="stat-value">{len(papers_dict)}</div>
                    <div class="stat-label">Papers</div>
                </div>
                <div class="stat">
                    <div class="stat-value">{total_citing_papers}</div>
                    <div class="stat-label">Citing Papers</div>
                </div>
                <div class="stat">
                    <div class="stat-value">{total_authors}</div>
                    <div class="stat-label">Total Authors</div>
                </div>
                <div class="stat">
                    <div class="stat-value">{authors_with_profiles}</div>
                    <div class="stat-label">With Profiles</div>
                </div>
            </div>
        </header>
        
        <div class="search-box">
            <input type="text" id="searchInput" placeholder="🔍 Search by name, affiliation, or paper title...">
        </div>
        
        <main id="content">
'''
    
    # Generate paper sections
    for paper_title, citing_papers in papers_dict.items():
        total_authors_in_paper = sum(len(authors) for authors in citing_papers.values())
        
        html += f'''
            <div class="paper-section" data-paper="{paper_title.lower()}">
                <div class="paper-header" onclick="togglePaper(this)">
                    <span class="paper-title">📄 {paper_title}</span>
                    <span class="paper-badge">{len(citing_papers)} citations · {total_authors_in_paper} authors</span>
                    <span class="toggle-icon">▼</span>
                </div>
                <div class="paper-content">
'''
        
        for citing_title, authors in citing_papers.items():
            html += f'''
                    <div class="citing-paper" data-citing="{citing_title.lower()}">
                        <div class="citing-title">📝 {citing_title}</div>
                        <div class="authors-grid">
'''
            
            for author in authors:
                profile_class = "" if author['has_profile'] else "no-profile"
                profile_icon = '<svg class="profile-icon" viewBox="0 0 24 24"><path d="M9 16.17L4.83 12l-1.42 1.41L9 19 21 7l-1.41-1.41z"/></svg>' if author['has_profile'] else ''
                
                name_html = f'<a href="{author["scholar_url"]}" target="_blank">{author["author_name"]}</a>' if author['scholar_url'] else author['author_name']
                
                affiliation = author['affiliation'] or 'No affiliation info'
                email = f"📧 {author['email_domain']}" if author['email_domain'] else ''
                
                html += f'''
                            <div class="author-card {profile_class}" data-name="{author['author_name'].lower()}" data-affiliation="{(author['affiliation'] or '').lower()}">
                                <div class="author-name">{profile_icon}{name_html}</div>
                                <div class="author-affiliation">{affiliation}</div>
                                {f'<div class="author-email">{email}</div>' if email else ''}
                            </div>
'''
            
            html += '''
                        </div>
                    </div>
'''
        
        html += '''
                </div>
            </div>
'''
    
    html += f'''
        </main>
        
        <footer>
            <p>Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>
            <p>Data sourced from Google Scholar</p>
        </footer>
    </div>
    
    <script>
        function togglePaper(header) {{
            const section = header.parentElement;
            section.classList.toggle('open');
        }}
        
        // Search functionality
        document.getElementById('searchInput').addEventListener('input', function(e) {{
            const query = e.target.value.toLowerCase().trim();
            const sections = document.querySelectorAll('.paper-section');
            
            if (!query) {{
                // Show all
                sections.forEach(s => {{
                    s.classList.remove('hidden');
                    s.querySelectorAll('.citing-paper').forEach(c => c.classList.remove('hidden'));
                    s.querySelectorAll('.author-card').forEach(a => a.classList.remove('hidden'));
                }});
                return;
            }}
            
            sections.forEach(section => {{
                let sectionMatch = false;
                const paperTitle = section.dataset.paper;
                
                if (paperTitle.includes(query)) {{
                    sectionMatch = true;
                }}
                
                section.querySelectorAll('.citing-paper').forEach(citing => {{
                    let citingMatch = citing.dataset.citing.includes(query);
                    
                    citing.querySelectorAll('.author-card').forEach(author => {{
                        const nameMatch = author.dataset.name.includes(query);
                        const affMatch = author.dataset.affiliation.includes(query);
                        
                        if (nameMatch || affMatch) {{
                            author.classList.remove('hidden');
                            citingMatch = true;
                        }} else if (!citingMatch) {{
                            author.classList.add('hidden');
                        }} else {{
                            author.classList.remove('hidden');
                        }}
                    }});
                    
                    if (citingMatch) {{
                        citing.classList.remove('hidden');
                        sectionMatch = true;
                    }} else {{
                        citing.classList.add('hidden');
                    }}
                }});
                
                if (sectionMatch) {{
                    section.classList.remove('hidden');
                    section.classList.add('open');
                }} else {{
                    section.classList.add('hidden');
                }}
            }});
        }});
        
        // Open first section by default
        document.querySelector('.paper-section')?.classList.add('open');
    </script>
</body>
</html>
'''
    
    filename = 'citations_network.html'
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"✅ HTML file saved: {filename}")
    return filename

def main():
    """Main function to run the citation extractor"""
    start_time = time.time()
    
    print("\n" + "🚀 Starting Citation Extractor...")
    print("This may take several minutes depending on the number of citations.\n")
    
    # Process all citations
    data, author_info = process_all_citations()
    
    if not data:
        print("\n❌ No citation data found!")
        return
    
    print(f"\n{'='*60}")
    print(f"📊 PROCESSING COMPLETE")
    print(f"{'='*60}")
    print(f"Total records: {len(data)}")
    
    # Generate outputs
    print("\n📁 Generating output files...")
    generate_xlsx(data, author_info.get('name', 'Scholar'))
    generate_html(data, author_info)
    
    elapsed = time.time() - start_time
    print(f"\n⏱️  Total time: {elapsed:.1f} seconds ({elapsed/60:.1f} minutes)")
    print("\n✨ Done! Open citations_network.html in your browser to explore the data.")

if __name__ == "__main__":
    main()

